from dotenv import load_dotenv
from config.db import connect_to_mysql, command, query
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, Bot
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, CallbackContext, ChatJoinRequestHandler, CallbackQueryHandler, PollAnswerHandler, ChatMemberHandler
import os
import pandas as pd

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import datetime
from utils import ordinal, plural_number
load_dotenv()
from stage import check_stage, upload_stage, finish_upload_stage
import re
import threading
from function import check_whitelist_user, get_all_groups, is_admin, insert_poll, get_poll_by_id
import traceback

# Aturan poin
MESSAGE_POINTS = 100
MEDIA_POINTS = 200
MAX_MESSAGE_POINTS = 700
MAX_MEDIA_POINTS = 400
REFERRAL_ACTIVE_DAYS = 10
REFERRED_MIN_ACTIVATION = 3
REFERRAL_POINTS = 200
MAX_LEADERBOARD_DATA_PER_PAGE = 5
MAX_REFERRAL_PER_DAY = 2
def register_user(user_id, username, first_name, last_name, group_id, context):
    # Register group if not exists
    data_group = query("SELECT group_id FROM `groups` WHERE group_id = %s", (group_id,), single=True)
    if(data_group) is None:
        bot = Bot(os.getenv("TELEGRAM_BOT_TOKEN"))
        chat = bot.get_chat(chat_id=group_id)
        group_name = chat.title
        command("INSERT INTO `groups` (group_id, group_name) VALUES (%s, %s)", (group_id, group_name))
    # Register user if not exists
    data = query("SELECT user_id FROM users WHERE user_id = %s", (user_id,), single=True)
    if(data) is None:
        result = command("INSERT INTO users (user_id, username, first_name, last_name) VALUES (%s, %s, %s, %s)", (user_id, username, first_name, last_name)) 
        if result is not None:
            current_date = datetime.datetime.now().strftime("%Y-%m-%d")
            data_user = query("SELECT id FROM users WHERE user_id = %s", (user_id,), single=True)
            print(data_user['id'], 0, 'registration', 100, current_date, group_id)
            if not is_admin(group_id, user_id, context):
                command("INSERT INTO scores (user_id, message_id, activity_type, score, date, group_id) VALUES (%s, %s, %s, %s, %s, %s)", (data_user['id'], 0, 'registration', 100, current_date, group_id))       
    else:
        data = query("SELECT users.user_id FROM users LEFT JOIN scores ON users.id = scores.user_id WHERE users.user_id = %s AND activity_type = 'registration' AND group_id = %s", (user_id, group_id), single=True)
        if data is None:
            current_date = datetime.datetime.now().strftime("%Y-%m-%d")
            data_user = query("SELECT id FROM users WHERE user_id = %s", (user_id,), single=True)
            if not is_admin(group_id, user_id, context):
                command("INSERT INTO scores (user_id, message_id, activity_type, score, date, group_id) VALUES (%s, %s, %s, %s, %s, %s)", (data_user['id'], 0, 'registration', 100, current_date, group_id))       
    return True

def get_daily_points(user_id, activity_type, group_id):
    """Hitung total poin harian pengguna berdasarkan jenis aktivitas."""
    data = query("SELECT COALESCE(SUM(score), 0) as score FROM scores LEFT JOIN users ON scores.user_id = users.id WHERE users.user_id = %s AND activity_type = %s AND DATE(`date`) = CURDATE() AND group_id = %s", (user_id, activity_type, group_id), single=True)
    print(data)
    return data['score']

def add_points(update: Update, user_id, message_id, group_id, activity_type, points, max_points, context: CallbackContext):
    user = update.message.from_user
    if user.is_bot:
        return
    current_points = int(get_daily_points(user_id, activity_type, update.message.chat_id))
    # Check if user is already referred
    referred = query("SELECT referrals.id, status, referred_id, referrer_id, referrals.created_at, group_id, DATE_FORMAT(DATE_ADD(referrals.created_at, INTERVAL 10 DAY), '%Y-%m-%d') as last_date FROM referrals LEFT JOIN referral_links ON referrals.link_id = referral_links.id WHERE referred_id = %s AND status = 0 AND group_id = %s", (user_id, group_id), single=True)
    print(referred)
    # For referral
    current_date_now = datetime.datetime.now()
    if referred is not None:
        id = referred['id']
        referrer_id = referred['referrer_id']
        referred_id = referred['referred_id']
        created_at = referred['created_at']
        date_now = datetime.datetime.now()
        date1 = date_now.strftime("%Y-%m-%d %H:%M:%S")
        first_date_of_refered = created_at
        last_date_of_referred = referred['last_date']
        referred_details = query("SELECT COUNT(referral_details.id) as total_data FROM referrals LEFT JOIN referral_details ON referrals.id = referral_details.referral_id WHERE referrals.id = %s AND `date` BETWEEN %s AND %s", (id, first_date_of_refered, last_date_of_referred), single=True)
        print(last_date_of_referred)
        data_referrer = query("SELECT id, username FROM users WHERE user_id = %s", (referrer_id,), single=True)
        date_obj = datetime.datetime.strptime(last_date_of_referred, "%Y-%m-%d")
        print(referred_details)
        if referred_details['total_data'] + 1 < REFERRED_MIN_ACTIVATION and date_now > date_obj:
            command("UPDATE referrals SET status = %s WHERE referrer_id = %s", (2, referrer_id))    
        elif referred_details['total_data'] + 1 < MAX_REFERRAL_PER_DAY:
            command("INSERT INTO referral_details (referral_id, `date`) VALUES(%s, %s)", (id, date1))
        else:
            print(f"GOT A REFERRAL FOR USER {data_referrer['username']}")
            insert_referral_detail = command("INSERT INTO referral_details (referral_id, `date`) VALUES(%s, %s)", (id, date1))
            max_referred_start_date = datetime.datetime.now().strftime("%Y-%m-%d") + " 00:00:00"
            max_referred_end_date = datetime.datetime.now().strftime("%Y-%m-%d") + " 23:59:59"
            max_referred = query("SELECT COUNT(referral_details.referral_id) as total_data FROM referrals LEFT JOIN referral_links ON referrals.link_id = referral_links.id LEFT JOIN referral_details ON referrals.id = referral_details.referral_id WHERE referred_id = %s AND group_id = %s AND referral_details.date BETWEEN %s AND %s", (referred_id, group_id, max_referred_start_date, max_referred_end_date), single=True) # Query Get total referred today

            print(max_referred['total_data'] / REFERRED_MIN_ACTIVATION)
            print(MAX_REFERRAL_PER_DAY)

            if max_referred['total_data'] / REFERRED_MIN_ACTIVATION > MAX_REFERRAL_PER_DAY:
                print(f"MAX REFFERAL REACHED FOR USER {data_referrer['username']}")

            elif referred_details['total_data'] + 1 >= REFERRED_MIN_ACTIVATION and insert_referral_detail is not None:
                command("UPDATE referrals SET status = %s WHERE referrer_id = %s", (1, referrer_id))
                if not is_admin(group_id, user_id, context):
                    print((data_referrer['id'], 0, 'referral', REFERRAL_POINTS, current_date_now.strftime("%Y-%m-%d %H:%M:%S"), group_id))
                    command("INSERT INTO scores (user_id, message_id, activity_type, score, date, group_id) VALUES (%s, %s, %s, %s, %s, %s)", (data_referrer['id'], 0, 'referral', REFERRAL_POINTS, current_date_now.strftime("%Y-%m-%d %H:%M:%S"), group_id))
                    update.message.reply_text(f"{data_referrer['username']} earned {REFERRAL_POINTS} points.")
                
    if current_points is not None:
        if current_points < max_points:
            new_points = min(points, max_points - current_points)
            current_date = current_date_now.strftime("%Y-%m-%d")
            
            if not is_admin(group_id, user_id, context):
                data_user = query("SELECT id FROM users WHERE user_id = %s", (user_id,), single=True)
                command("INSERT INTO scores (user_id, message_id, activity_type, score, date, group_id) VALUES (%s, %s, %s, %s, %s, %s)", (data_user['id'], message_id, activity_type, new_points, current_date, group_id))

def process_upload_points(update: Update, context: CallbackContext):
    # Process uploaded file
    try:
        document = update.message.document
        file_path = f"./uploads"
        if not os.path.exists(file_path):
            os.mkdir(file_path)
        user = update.message.from_user
        now = datetime.datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
        uploaded_filename = f"uploaded_{user.id}_{update.message.chat_id}{now}.xlsx, xls"
        file_path = os.path.join(file_path, uploaded_filename)
        document.get_file().download(file_path)
        df = pd.read_excel(file_path)
        usernames = df["username"]
        points = df["points"]
        group_id = context.user_data['group_id']
        for username, point in zip(usernames, points): 
            username = re.sub(r"^@\S+\s*", "", username)
            current_date = datetime.datetime.now().strftime("%Y-%m-%d")
            data_user = query("SELECT id FROM users WHERE username = %s", (username,), single=True)
            if data_user is not None:
                command("INSERT INTO scores (user_id, message_id, activity_type, score, date, group_id) VALUES (%s, %s, %s, %s, %s, %s)", (data_user['id'], 0, 'extra point', point, current_date, group_id))
        os.remove(file_path)
        return True
    except Exception as e:
        traceback.print_exc()
        update.message.reply_text(f"Error processing uploaded file. File format is not valid. Try use the excel from /upload_points_template")
        return False
        

def poll_answer_handler(update: Update, context: CallbackContext):
    user_id = update.poll_answer.user.id
    poll_id = update.poll_answer.poll_id
    user = update.poll_answer.user
    if user.is_bot:
        return
    
    
    data = get_poll_by_id(poll_id)
    if data is not None:
        group_id = data['group_id']
        if group_id < 0:
            add_points(update, user_id, 0, group_id, "poll", MESSAGE_POINTS, MAX_MESSAGE_POINTS, context)

def poll_post_handler(update: Update, context: CallbackContext):
    if update.message and update.message.poll:
        poll_id = update.message.poll.id
        group_id = update.message.chat_id
        chat_id = update.message.chat.id
        insert_poll(poll_id, group_id, chat_id)
        # polls[poll_id] = chat_id
        

def handle_message(update: Update, context: CallbackContext):
    # Add points when users send a message in group
    user = update.message.from_user
    chat_id = update.message.chat_id  # ID Group or Chat
    message_id = update.message.message_id
    if user.is_bot:
        return

    
    if chat_id > 0:
        is_whitelist = check_whitelist_user(update.message.from_user.id)
        if not is_whitelist:
            update.message.reply_text("You are not authorized to use this bot.")
            return
    else:
        res = register_user(user.id, user.username, user.first_name, user.last_name, chat_id, context)
        if res is None:
            print("Failed to register user!")
            return
    if update.message.photo or update.message.video or update.message.animation or update.message.document or update.message.sticker:
        current_stage = check_stage(update.message.from_user.id)
        if current_stage == 1 and update.message.document and chat_id > 0:
            document = update.message.document
        
            # Get file name and MIME type
            file_name = document.file_name
            mime_type = document.mime_type
            # Check file extension
            if not file_name.endswith((".xlsx", ".xls")):
                update.message.reply_text(f"File '{file_name}' is not an Excel file based on its extension.")
                return

            if not mime_type in ["application/vnd.ms-excel", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
                update.message.reply_text(f"File '{file_name}' is not an Excel file based on its MIME type.")
                return

            # Upload file
            thread = threading.Thread(target=process_upload_points, args=(update, context))
            thread.start()
            update.message.reply_text(f"File '{file_name}' is succesfully uploaded use /finish_upload to finish the upload process.")
            # if res == True:
            #     update.message.reply_text(f"File '{file_name}' is succesfully uploaded")

        elif current_stage != 1 or chat_id < 0:
            add_points(update, user.id, message_id, chat_id, "media", MEDIA_POINTS, MAX_MEDIA_POINTS, context)
        else:
            update.message.reply_text("The uploaded file is not an Excel file.")
            return
    else:
        if check_stage(update.message.from_user.id) == 1 and chat_id > 0:
            update.message.reply_text("You are in the process of uploading points with excel file. Please upload xlsx, xls file format only. Use /finish_upload to cancel the process.")
            return
        elif chat_id < 0:
            add_points(update, user.id, message_id, chat_id, "message", MESSAGE_POINTS, MAX_MESSAGE_POINTS, context)

def myscore(update: Update, context: CallbackContext):
    user = update.message.from_user
    if user.is_bot:
        return
    # Command to show users score
    chat_id = update.message.chat_id
    if chat_id > 0:
        return 
    register_user(update.message.from_user.id, update.message.from_user.username, update.message.from_user.first_name, update.message.from_user.last_name, update.message.chat_id, context)
    user_id = update.message.from_user.id
    data = query("SELECT COALESCE(SUM(score), 0) as score, first_name, last_name FROM scores LEFT JOIN users ON scores.user_id = users.id WHERE users.user_id = %s AND group_id = %s", (user_id, chat_id), single=True)
    
    data_leaderboard = query(
        "SELECT u.username, COALESCE(SUM(s.score), 0) as total_points, u.user_id FROM users u "
        "LEFT JOIN scores s ON u.id = s.user_id "
        " WHERE group_id = %s"
        " GROUP BY u.user_id ORDER BY total_points DESC", params=(update.message.chat_id, ), single=False
    )

    rank = 0
    for i, row in enumerate(data_leaderboard, start=1):
        if row['user_id'] == user_id:
            rank = i

    total_score = data['score']
    msg = f"Your score: {plural_number(total_score, 'point')}\n\nRank: {ordinal(rank)}"
    update.message.reply_text(msg)

def leaderboard(update: Update, context: CallbackContext):
    page = 1
    max_index = 5
    leaderboard_text = "🏆 Leaderboard 🏆\n\n"
    q = update.callback_query
    if q is None:
        chat_id = update.message.chat_id
        data = query(
            "SELECT u.username, COALESCE(SUM(s.score), 0) as total_points, `date` FROM users u "
            "LEFT JOIN scores s ON u.id = s.user_id"
            " WHERE group_id = %s"
            " GROUP BY u.user_id ORDER BY total_points DESC, `date` DESC", params=(chat_id,), single=False
        )
        if data is None:
            update.message.reply_text("Leaderboard is empty.")
            return
        elif len(data) == 0:
            update.message.reply_text("Leaderboard is empty.")
            return
        leaderboard_users = []
        for i, row in enumerate(data, start=1):
            total_point = row['total_points']
            username = row['username'] if row['username'] else "Unknown"
            leaderboard_users.append(f"{i}. {username} - {total_point} points\n")
        cutted_data = leaderboard_users[:max_index]
        leaderboard_text += "".join(cutted_data)
        if len(leaderboard_users) > MAX_LEADERBOARD_DATA_PER_PAGE:
            keyboard = InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton("Load More", callback_data=f"callback_leaderboard_{page + 1}"),
                    ]
                ]
            )
            # leaderboard_text += f"\nShow users {(page) * len(cutted_data)} of {len(leaderboard_users)}"
            update.message.reply_text(leaderboard_text, reply_markup=keyboard)
        else:
            # leaderboard_text += f"\nShow users {len(leaderboard_users)} of {len(leaderboard_users)}"
            update.message.reply_text(leaderboard_text)
    else:
        q.answer()
        q_data = q.data
        q_data = q_data.replace("callback_leaderboard_", "")
        page = int(q_data)
        start_index = 5 + (page - 2) * 10
        end_index = start_index + 10
        chat_id = update.callback_query.message.chat_id
        data = query(
            "SELECT u.username, COALESCE(SUM(s.score), 0) as total_points, `date` FROM users u "
            "LEFT JOIN scores s ON u.id = s.user_id"
            " WHERE group_id = %s"
            " GROUP BY u.user_id ORDER BY total_points DESC, `date` DESC", params=(chat_id,), single=False
        )
        if data is None:
            update.message.reply_text("Leaderboard is empty.")
            return
        elif len(data) == 0:
            update.message.reply_text("Leaderboard is empty.")
            return
        leaderboard_users = []
        for i, row in enumerate(data, start=1):
            total_point = row['total_points']
            username = row['username'] if row['username'] else "Unknown"
            leaderboard_users.append(f"{i}. {username} - {total_point} points\n")
        cutted_data = leaderboard_users[start_index:end_index]
        leaderboard_text += "".join(cutted_data)
        if end_index >= len(leaderboard_users):
            # leaderboard_text += f"\nShow users {len(leaderboard_users)} of {len(leaderboard_users)}"
            q.edit_message_reply_markup(reply_markup=None)
            context.bot.send_message(chat_id=chat_id, text=leaderboard_text)
        else:
            keyboard = InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton("Load More", callback_data=f"callback_leaderboard_{page + 1}"),
                    ]
                ]
            )
            # leaderboard_text += f"\nShow users {(page) * len(cutted_data)} of {(page) * len(cutted_data)}"
            q.edit_message_reply_markup(reply_markup=None)
            context.bot.send_message(chat_id=chat_id, text=leaderboard_text)
            # keyboard = InlineKeyboardMarkup(
            #     [
            #         [
            #             InlineKeyboardButton("Load More", callback_data=f"callback_leaderboard_{page + 1}"),
            #         ]
            #     ]
            # )
            # leaderboard_text += f"\nShow users {(page) * len(cutted_data)} of {len(leaderboard_users)}"
            # q.edit_message_reply_markup(reply_markup=None)
            # context.bot.send_message(chat_id=chat_id, text=leaderboard_text, reply_markup=keyboard)
def handle_start(update: Update, context: CallbackContext):
    print("LAH")
    chat_id = update.message.chat_id
    user = update.message.from_user
    if user.is_bot:
        return
    if chat_id > 0:
        is_whitelist = check_whitelist_user(update.message.from_user.id)
        if not is_whitelist:
            update.message.reply_text("You are not authorized to use this bot.")
            return
        if check_stage(update.message.from_user.id) == 1:
            update.message.reply_text("You are in the process of uploading points with excel file. Please upload xlsx, xls file format only. Use /finish_upload to cancel the process.")
            return
        update.message.reply_text("Hello! 👋🏽\n\nTelegram Bot Commands: \n\nExport Score - /export_scores\nUpload Points - /upload_points\nDownload Upload Points Template /upload_points_template")
        return 
    
    register_user(update.message.from_user.id, update.message.from_user.username, update.message.from_user.first_name, update.message.from_user.last_name, update.message.chat_id, context)
    msg = f"Hello {update.message.from_user.username} 👋🏽\n\nEngagement Tracking Commands: \n\nCheck Progress - /myscore🥇\nLeaderboard - /leaderboard 📊\nInvite Friends - /create_referral👥"
    update.message.reply_text(msg)

def upload_points_template(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    if chat_id < 0:
        return 
    
    is_whitelist = check_whitelist_user(update.message.from_user.id)
    if not is_whitelist:
        update.message.reply_text("You are not authorized to use this bot.")
        return

    if check_stage(update.message.from_user.id) == 1:
        update.message.reply_text("You are in the process of uploading points with excel file. Please upload xlsx, xls file format only. Use /finish_upload to cancel the process.")
        return
    # register_user(update.message.from_user.id, update.message.from_user.username, update.message.from_user.first_name, update.message.from_user.last_name, update.message.chat_id)
    with open('excel_template/template_upload_points.xlsx', 'rb') as excel_file:
        context.bot.send_document(
            chat_id = chat_id,
            document=excel_file,
            filename="upload_points_template.xlsx",
            caption="Upload Points Template"
        )

def export_scores(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    if chat_id > 0:
        is_whitelist = check_whitelist_user(update.message.from_user.id)
        if not is_whitelist:
            update.message.reply_text("You are not authorized to use this bot.")
            return
        if check_stage(update.message.from_user.id) == 1:
            update.message.reply_text("You are in the process of uploading points with excel file. Please upload xlsx, xls file format only. Use /finish_upload to cancel the process.")
            return
    
        # register_user(update.message.from_user.id, update.message.from_user.username, update.message.from_user.first_name, update.message.from_user.last_name, update.message.chat.id)
        groups = get_all_groups()
        if groups is None:
            update.message.reply_text("No groups found!")
            return
        if len(groups) > 1:
            keyboard = InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(group['group_name'], callback_data=f"callback_export_scores_{group['group_id']}")
                    ] for group in groups
                ]
            )
            update.message.reply_text("Select a group to export scores:", reply_markup=keyboard)
        else:
            handle_export_scores(update, context, groups[0]['group_id'])
        return 

def handle_export_scores(update: Update, context: CallbackContext, group_id):
    q = update.callback_query
    if q is not None:
        q.answer()
    chat_id = group_id
    if chat_id > 0:
        return 
    date = datetime.datetime.now().strftime("%Y-%m-%d")
    data = query(
        "SELECT u.username, first_name, last_name, COALESCE(SUM(s.score), 0) as total_points FROM users u "
        "LEFT JOIN scores s ON u.id = s.user_id "
        " WHERE group_id = %s"
        " GROUP BY u.user_id ORDER BY total_points DESC, `date` DESC", dictionary=False, params=(group_id,), single=False
    )

    if(not data):
        if q is not None:
            context.bot.delete_message(chat_id=q.message.chat_id, message_id=q.message.message_id)
            context.bot.send_message(chat_id=chat_id, text="No data found!")
        else:
            update.message.reply_text("No data found!")
        return
    
    date = datetime.datetime.now().strftime("%Y-%m-%d")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"User Scores {date}"
    headers = ["Username", "First Name", "Last Name", "Score"]

    ws.append(headers)

    # Add data to worksheet
    for row in data:
        ws.append(row)


    # Buat border untuk semua sel
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border  # Add borders to all cells

    # Atur lebar kolom otomatis
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # get all column letters
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # add margin


    if not os.path.exists("export_xls"):
        os.mkdir("export_xls")
    date = datetime.datetime.now().strftime("%Y-%m-%d")
    filename = f"export_xls/score_{chat_id}_{date}.xlsx"
    # Save the excel file
    wb.save(filename)

    if q is not None:
        context.bot.delete_message(chat_id=q.message.chat_id, message_id=q.message.message_id)
        context.bot.send_document(chat_id=q.message.chat_id, document=open(filename, 'rb'))
    else:
        update.message.reply_document(document=open(filename, 'rb'))

def create_referral(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    if chat_id > 0:
        return 
    
    user = update.message.from_user
    if user.is_bot:
        return
    
    register_user(update.message.from_user.id, update.message.from_user.username, update.message.from_user.first_name, update.message.from_user.last_name, update.message.chat_id, context)
    invite_link = context.bot.create_chat_invite_link(chat_id=update.message.chat_id, creates_join_request=True)
    referral_message = f"Here is your referral link: {invite_link.invite_link}\n"
    # current_date = datetime.datetime.now()
    data = query("SELECT id FROM users WHERE user_id = %s", (update.message.from_user.id,), single=True)
    if command("INSERT INTO referral_links (user_id, link, group_id) VALUES (%s, %s, %s)", (data['id'], invite_link.invite_link, update.message.chat_id)) is not None:
        update.message.reply_text(referral_message)
    else:
        update.message.reply_text("Failed to create referral link.")

def handle_join_request(update: Update, context: CallbackContext):
    join_request = update.chat_join_request
    user = join_request.from_user
    invite_link = join_request.invite_link.invite_link
    data = query("SELECT referral_links.id, users.user_id FROM referral_links LEFT JOIN users ON referral_links.user_id = users.id WHERE link = %s and group_id = %s", (invite_link, join_request.chat.id), single=True)
    register_user(join_request.from_user.id, join_request.from_user.username, join_request.from_user.first_name, join_request.from_user.last_name, join_request.chat.id, context)
    if data is not None:
        check_another_referral = query("SELECT COUNT(referrals.id) as total_data FROM referrals LEFT JOIN referral_links ON referrals.link_id = referral_links.id WHERE referred_id = %s and group_id = %s", (user.id, join_request.chat.id), single=True)
        if check_another_referral['total_data'] == 0:
            command("INSERT INTO referrals (referrer_id, referred_id, link_id) VALUES (%s, %s, %s)", (data['user_id'], user.id, data['id'])) # Join with referrer link
    join_request.approve()

def upload_points(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    if chat_id > 0:
        is_whitelist = check_whitelist_user(update.message.from_user.id)
        if not is_whitelist:
            update.message.reply_text("You are not authorized to use this bot.")
            return
        if check_stage(update.message.from_user.id) == 1:
            update.message.reply_text("You are in the process of uploading points with excel file. Please upload xlsx, xls file format only. Use /finish_upload to cancel the process.")
            return
    
        # register_user(update.message.from_user.id, update.message.from_user.username, update.message.from_user.first_name, update.message.from_user.last_name, update.message.chat.id)
        groups = get_all_groups()
        if groups is None:
            update.message.reply_text("No groups found!")
            return
        if len(groups) > 1:
            keyboard = InlineKeyboardMarkup(
                [
                    [
                        InlineKeyboardButton(group['group_name'], callback_data=f"callback_upload_points_{group['group_id']}")
                    ] for group in groups
                ]
            )
            update.message.reply_text("Select a group to upload points:", reply_markup=keyboard)
        else:
            handle_upload_points(update, context, groups[0]['group_id'])
        return 

def handle_upload_points(update: Update, context: CallbackContext, group_id):
    q = update.callback_query
    if q is not None:
        q.answer()
        chat_id = q.message.chat_id
    else:
        chat_id = update.message.from_user.id
    context.user_data['group_id'] = group_id
    if check_stage(chat_id) == 1:
        if q is not None:
            context.bot.delete_message(chat_id=chat_id, message_id=q.message.message_id)
            context.bot.send_message(chat_id=q.message.chat_id, text="You are in the process of uploading points with excel file. Please upload xlsx, xls file format only. Use /finish_upload to cancel the process.")
        else:
            update.message.reply_text("You are in the process of uploading points with excel file. Please upload xlsx, xls file format only. Use /finish_upload to cancel the process.")
        return
    else:
        upload_stage(chat_id)
        if q is not None:
            context.bot.delete_message(chat_id=chat_id, message_id=q.message.message_id)
            context.bot.send_message(chat_id=q.message.chat_id, text="Please upload xlsx, xls file format only to add points.")
        else:
            update.message.reply_text("Please upload xlsx, xls file format only to add points.")
        return


def handle_query_callback(update: Update, context: CallbackContext):
    q = update.callback_query
    q.answer()
    if q.data.startswith("callback_leaderboard_"):
        leaderboard(update, context)
        return
    elif q.data.startswith("callback_export_scores_"):
        handle_export_scores(update, context, int(q.data.replace("callback_export_scores_", "")))
        return
    elif q.data.startswith("callback_upload_points_"):
        handle_upload_points(update, context, int(q.data.replace("callback_upload_points_", "")))
        return

def finish_upload(update: Update, context: CallbackContext):
    chat_id = update.message.chat_id
    if chat_id < 0:
        return
    
    if check_stage(chat_id) == 1:
        user = update.message.from_user
        update.message.reply_text("Upload file with excel format has been finished.")
        finish_upload_stage(user.id)
        context.user_data.clear()
        return
    else:
        update.message.reply_text("You're not in the process of uploading points with excel file.")

# def welcome(update: Update, context: CallbackContext):
#     print(f"WELCOME {update.message.from_user.username}")
#     register_user(update.message.from_user.id, update.message.from_user.username, update.message.from_user.first_name, update.message.from_user.last_name, update.message.chat_id, context)

def welcome2(update: Update, context: CallbackContext):
    chat_member = update.chat_member
    if not chat_member:  # Prevents 'NoneType' errors
        return

    # Get the new member's status
    new_status = chat_member.new_chat_member.status
    print(new_status)
    # Ensure the user actually joined
    if new_status == "member":
        chat = chat_member.chat
        user = chat_member.new_chat_member.user
        print("INSERT NEW MEMBER")
        print(user.id, user.username, user.first_name, user.last_name, chat.id)
        register_user(user.id, user.username, user.first_name, user.last_name, chat.id, context)
        # context.bot.send_message(chat_id=chat.id, text=f"WELCOME {user.full_name}!")

def main():
    global MESSAGE_POINTS, MEDIA_POINTS, MAX_MESSAGE_POINTS, MAX_MEDIA_POINTS, REFERRAL_ACTIVE_DAYS, REFERRED_MIN_ACTIVATION, REFERRAL_POINTS, MAX_LEADERBOARD_DATA_PER_PAGE, MAX_REFERRAL_PER_DAY
    updater = Updater(token=os.getenv("TELEGRAM_BOT_TOKEN"), use_context=True)
    dp = updater.dispatcher

    bot_config = query("SELECT * FROM bot_config", dictionary=True, single=True)
    if(bot_config):
        MESSAGE_POINTS = int(bot_config['MESSAGE_POINTS'])
        MEDIA_POINTS = int(bot_config['MEDIA_POINTS'])
        MAX_MESSAGE_POINTS = int(bot_config['MAX_MESSAGE_POINTS'])
        MAX_MEDIA_POINTS = int(bot_config['MAX_MEDIA_POINTS'])
        REFERRAL_ACTIVE_DAYS = int(bot_config['REFERRAL_ACTIVE_DAYS'])
        REFERRED_MIN_ACTIVATION = int(bot_config['REFERRED_MIN_ACTIVATION'])
        REFERRAL_POINTS = int(bot_config['REFERRAL_POINTS'])
        MAX_LEADERBOARD_DATA_PER_PAGE = int(bot_config['MAX_LEADERBOARD_DATA_PER_PAGE'])
        MAX_REFERRAL_PER_DAY = int(bot_config['MAX_REFERRAL_PER_DAY'])

    dp.add_handler(CommandHandler("start", handle_start))
    dp.add_handler(CommandHandler("myscore", myscore))
    dp.add_handler(CommandHandler("leaderboard", leaderboard))
    dp.add_handler(CommandHandler("export_scores", export_scores))
    dp.add_handler(CommandHandler("create_referral", create_referral))
    dp.add_handler(CommandHandler("finish_upload", finish_upload))
    dp.add_handler(CommandHandler("upload_points", upload_points))
    dp.add_handler(CommandHandler("upload_points_template", upload_points_template))
    dp.add_handler(CallbackQueryHandler(handle_query_callback))
    dp.add_handler(ChatJoinRequestHandler(handle_join_request))
    dp.add_handler(PollAnswerHandler(poll_answer_handler))
    dp.add_handler(MessageHandler(Filters.text | Filters.photo | Filters.video | Filters.animation | Filters.document | Filters.sticker, handle_message))
    dp.add_handler(MessageHandler(Filters.poll, poll_post_handler))
    # dp.add_handler(MessageHandler(Filters.status_update.new_chat_members, welcome))
    dp.add_handler(ChatMemberHandler(welcome2, ChatMemberHandler.ANY_CHAT_MEMBER))
    print("Running bot!")
    updater.start_polling(allowed_updates=Update.ALL_TYPES)
    updater.idle()


if __name__ == "__main__":
    main()